import os
import re
import sys
from collections import defaultdict
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

CURRENT_DIR = os.path.abspath(os.path.dirname(__file__))
BACKEND_DIR = os.path.dirname(CURRENT_DIR)
PROJECT_ROOT = os.path.dirname(BACKEND_DIR)

if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)
from backend.models import (
    DailyMachinePlan,
    DailyProductTonnage,
    Machine,
    MonthlyEfficiencySummary,
    MonthlyMachineSummary,
    MonthlyProductTonnage,
    Plant,
    Product,
)
from backend.models.db import db


RAW_FILES = {
    "daily": Path("/home/utba/flaskapp/ACI/Daily Total Tonnage 2026(1).xlsx"),
    "monthly": Path("/home/utba/flaskapp/ACI/Total Monthly Tonnage 2026(1).xlsx"),
    "plan": Path("/home/utba/flaskapp/ACI/PLAN 2026(1).xlsx"),
}

PLANTS = [
    {"code": "ACI", "name": "ACI / Suasa Plant"},
    {"code": "RS", "name": "RS / Keluli Plant"},
]

MACHINES = [
    {"plant_code": "ACI", "code": "530A", "name": "530A", "machine_group": "Group A", "display_order": 10},
    {"plant_code": "ACI", "code": "530B", "name": "530B", "machine_group": "Group A", "display_order": 20},
    {"plant_code": "ACI", "code": "570", "name": "570", "machine_group": "Group A", "display_order": 30},
    {"plant_code": "ACI", "code": "M8", "name": "M8", "machine_group": "Group A", "display_order": 40},
    {"plant_code": "ACI", "code": "MPL", "name": "MPL", "machine_group": "Group A", "display_order": 50},
    {"plant_code": "ACI", "code": "SPL", "name": "SPL", "machine_group": "Group A", "display_order": 60},
    {"plant_code": "RS", "code": "530C", "name": "530C", "machine_group": "Group B", "display_order": 70},
    {"plant_code": "RS", "code": "M16", "name": "M16", "machine_group": "Group B", "display_order": 80},
    {"plant_code": "RS", "code": "MANUAL", "name": "Manual Keluli", "machine_group": "Group B", "display_order": 90},
]

MONTH_NAME_MAP = {
    "JAN": 1,
    "JANUARY": 1,
    "FEB": 2,
    "FEBRUARY": 2,
    "MAR": 3,
    "MARCH": 3,
    "MAC": 3,
    "APR": 4,
    "APRIL": 4,
    "MAY": 5,
    "JUNE": 6,
    "JUN": 6,
    "JULY": 7,
    "JUL": 7,
    "AUG": 8,
    "AUGUST": 8,
    "SEP": 9,
    "SEPT": 9,
    "SEPTEMBER": 9,
    "OCT": 10,
    "OCTOBER": 10,
    "NOV": 11,
    "NOVEMBER": 11,
    "DEC": 12,
    "DECEMBER": 12,
}

MONTHLY_MACHINE_ROW_MAP = {
    4: "530A",
    5: "530B",
    6: "570",
    7: "M8",
    8: "MPL",
    9: "SPL",
    11: "M16",
    12: "530C",
    13: "MANUAL",
}

PLAN_MACHINE_ALIASES = {
    "530A": "530A",
    "530B": "530B",
    "570": "570",
    "M8": "M8",
    "MPL": "MPL",
    "MPL (MANUAL SUASA)": "MPL",
    "SPL": "SPL",
    "SPL (SMALL PACK SUASA)": "SPL",
    "530C": "530C",
    "N530": "530C",
    "M16": "M16",
    "MANUAL": "MANUAL",
    "MANUAL KELULI": "MANUAL",
    "MPL (MANUAL KELULI)": "MANUAL",
    "MANUAL BAGGING": "MANUAL",
}


def clean_string(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_decimal(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        text = value.strip()
        if not text or text == "#DIV/0!":
            return None
        try:
            return Decimal(text)
        except InvalidOperation:
            return None
    return None


def extract_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text or text == "#DIV/0!":
        return None
    if text.startswith("="):
        return None
    match = re.search(r"-?\\d+(?:\\.\\d+)?", text)
    if not match:
        return None
    return Decimal(match.group(0))


def normalize_machine_code(raw_value):
    label = clean_string(raw_value)
    if not label:
        return None
    normalized = " ".join(label.upper().replace("_", " ").split())
    return PLAN_MACHINE_ALIASES.get(normalized)


def ensure_master_data():
    plant_map = {}
    for payload in PLANTS:
        plant = Plant.query.filter_by(code=payload["code"]).first()
        if not plant:
            plant = Plant(code=payload["code"], name=payload["name"])
            db.session.add(plant)
            db.session.flush()
        else:
            plant.name = payload["name"]
            plant.is_active = True
        plant_map[payload["code"]] = plant

    machine_map = {}
    for payload in MACHINES:
        machine = Machine.query.filter_by(code=payload["code"]).first()
        if not machine:
            machine = Machine(
                plant_id=plant_map[payload["plant_code"]].id,
                code=payload["code"],
                name=payload["name"],
                machine_group=payload["machine_group"],
                display_order=payload["display_order"],
            )
            db.session.add(machine)
            db.session.flush()
        else:
            machine.plant_id = plant_map[payload["plant_code"]].id
            machine.name = payload["name"]
            machine.machine_group = payload["machine_group"]
            machine.display_order = payload["display_order"]
            machine.is_active = True
        machine_map[payload["code"]] = machine

    db.session.flush()
    return plant_map, machine_map


def clear_raw_tables():
    db.session.query(DailyProductTonnage).delete()
    db.session.query(MonthlyProductTonnage).delete()
    db.session.query(DailyMachinePlan).delete()
    db.session.query(MonthlyMachineSummary).delete()
    db.session.query(MonthlyEfficiencySummary).delete()
    db.session.query(Product).delete()
    db.session.flush()


def upsert_product(product_cache, plant_id, part_code, description, product_class, warehouse_code):
    cache_key = (part_code, warehouse_code)
    product = product_cache.get(cache_key)
    if not product:
        product = Product(
            plant_id=plant_id,
            part_code=part_code,
            description=description,
            product_class=product_class,
            warehouse_code=warehouse_code,
            is_active=True,
        )
        db.session.add(product)
        db.session.flush()
        product_cache[cache_key] = product
    else:
        product.plant_id = plant_id
        product.description = description
        product.product_class = product_class
        product.warehouse_code = warehouse_code
        product.is_active = True
    return product


def import_daily_product_tonnage(workbook_path: Path, plant_map, product_cache):
    workbook = load_workbook(workbook_path, data_only=True)
    imported = 0
    daily_totals = defaultdict(lambda: Decimal("0"))
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        month = MONTH_NAME_MAP.get(sheet_name.strip().upper())
        if not month:
            continue
        days_in_month = monthrange(2026, month)[1]

        for row in range(5, ws.max_row + 1):
            part_code = clean_string(ws.cell(row, 2).value)
            description = clean_string(ws.cell(row, 3).value)
            if not part_code or not description or not part_code.startswith("FG"):
                continue

            product = upsert_product(
                product_cache=product_cache,
                plant_id=plant_map["ACI"].id,
                part_code=part_code,
                description=description,
                product_class=clean_string(ws.cell(row, 4).value),
                warehouse_code=clean_string(ws.cell(row, 5).value),
            )

            for column in range(6, 37):
                day_value = ws.cell(3, column).value
                if not isinstance(day_value, int):
                    continue
                if day_value < 1 or day_value > days_in_month:
                    continue
                output_mt = to_decimal(ws.cell(row, column).value)
                if output_mt is None:
                    continue
                daily_totals[(date(2026, month, day_value), product.id)] += output_mt

    for (record_date, product_id), output_mt in daily_totals.items():
        db.session.add(
            DailyProductTonnage(
                record_date=record_date,
                product_id=product_id,
                output_mt=output_mt,
            )
        )
        imported += 1
    return imported


def import_monthly_product_tonnage(workbook_path: Path, plant_map, product_cache):
    workbook = load_workbook(workbook_path, data_only=True)
    ws = workbook["MONTHLY"]
    imported = 0
    monthly_totals = defaultdict(lambda: Decimal("0"))

    month_columns = {}
    for column in range(6, 18):
        header = clean_string(ws.cell(3, column).value)
        if not header:
            continue
        month = MONTH_NAME_MAP.get(header.upper())
        if month:
            month_columns[column] = month

    for row in range(5, ws.max_row + 1):
        part_code = clean_string(ws.cell(row, 2).value)
        description = clean_string(ws.cell(row, 3).value)
        if not part_code or not description or not part_code.startswith("FG"):
            continue

        product = upsert_product(
            product_cache=product_cache,
            plant_id=plant_map["ACI"].id,
            part_code=part_code,
            description=description,
            product_class=clean_string(ws.cell(row, 4).value),
            warehouse_code=clean_string(ws.cell(row, 5).value),
        )

        for column, month in month_columns.items():
            output_mt = to_decimal(ws.cell(row, column).value)
            if output_mt is None:
                continue
            monthly_totals[(date(2026, month, 1), product.id)] += output_mt

    for (month_start, product_id), output_mt in monthly_totals.items():
        db.session.add(
            MonthlyProductTonnage(
                month_start=month_start,
                product_id=product_id,
                output_mt=output_mt,
            )
        )
        imported += 1
    return imported


def import_monthly_machine_summary(workbook_path: Path, machine_map):
    workbook = load_workbook(workbook_path, data_only=True)
    product_class_ws = workbook["PRODUCT CLASS"]
    imported = 0
    month_columns = {}
    for column in range(3, 15):
        header = product_class_ws.cell(3, column).value
        if isinstance(header, datetime):
            month_columns[column] = header.date().replace(day=1)

    for row, machine_code in MONTHLY_MACHINE_ROW_MAP.items():
        machine = machine_map[machine_code]
        for column, month_start in month_columns.items():
            actual_output_mt = to_decimal(product_class_ws.cell(row, column).value)
            if actual_output_mt is None:
                continue
            db.session.add(
                MonthlyMachineSummary(
                    month_start=month_start,
                    machine_id=machine.id,
                    actual_output_mt=actual_output_mt,
                    rejected_output_mt=Decimal("0"),
                )
            )
            imported += 1

    return imported


def import_daily_machine_plan(workbook_path: Path, machine_map):
    workbook = load_workbook(workbook_path, data_only=True)
    imported = 0
    plan_totals = defaultdict(lambda: Decimal("0"))
    standard_output_by_key = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name == "Planning Efficiency":
            continue
        ws = workbook[sheet_name]
        expected_month = MONTH_NAME_MAP.get(sheet_name.replace("26", "").strip().upper())

        date_columns = {}
        for column in range(3, ws.max_column + 1):
            header = ws.cell(1, column).value
            if (
                isinstance(header, datetime)
                and header.year == 2026
                and (expected_month is None or header.month == expected_month)
            ):
                date_columns[column] = header.date()

        if not date_columns:
            continue

        for row in range(3, 12):
            machine_code = normalize_machine_code(ws.cell(row, 1).value)
            if not machine_code:
                continue

            machine = machine_map[machine_code]
            standard_output_mt = extract_number(ws.cell(row, 2).value)

            for column, plan_date in date_columns.items():
                planned_output_mt = to_decimal(ws.cell(row, column).value)
                if planned_output_mt is None:
                    continue
                plan_totals[(plan_date, machine.id)] += planned_output_mt
                if standard_output_mt is not None:
                    standard_output_by_key[(plan_date, machine.id)] = standard_output_mt

    for (plan_date, machine_id), planned_output_mt in plan_totals.items():
        db.session.add(
            DailyMachinePlan(
                plan_date=plan_date,
                machine_id=machine_id,
                planned_output_mt=planned_output_mt,
                standard_output_mt=standard_output_by_key.get((plan_date, machine_id)),
            )
        )
        imported += 1
    return imported


def import_monthly_efficiency_summary(workbook_path: Path):
    workbook = load_workbook(workbook_path, data_only=True)
    ws = workbook["Planning Efficiency"]
    imported = 0
    summary_rows = {}

    row = 1
    while row <= ws.max_row - 3:
        maybe_dates = ws.cell(row, 2).value
        planned_label = clean_string(ws.cell(row + 1, 1).value)
        output_label = clean_string(ws.cell(row + 2, 1).value)
        efficiency_label = clean_string(ws.cell(row + 3, 1).value)

        if not isinstance(maybe_dates, datetime) or planned_label != "Planned" or output_label != "Output" or efficiency_label != "Efficiency":
            row += 1
            continue

        scope_label = clean_string(ws.cell(row, 1).value) or "OVERALL"
        scope_code = normalize_machine_code(scope_label) or scope_label.upper().replace(" ", "_")

        for column in range(2, 15):
            month_start = ws.cell(row, column).value
            if not isinstance(month_start, datetime):
                continue
            planned_output_mt = to_decimal(ws.cell(row + 1, column).value)
            actual_output_mt = to_decimal(ws.cell(row + 2, column).value)
            efficiency_ratio = to_decimal(ws.cell(row + 3, column).value)
            if planned_output_mt is None and actual_output_mt is None and efficiency_ratio is None:
                continue
            summary_rows[(scope_code, month_start.date().replace(day=1))] = {
                "planned_output_mt": planned_output_mt or Decimal("0"),
                "actual_output_mt": actual_output_mt or Decimal("0"),
                "efficiency_ratio": efficiency_ratio,
            }
        row += 4

    for (scope_code, month_start), payload in summary_rows.items():
        db.session.add(
            MonthlyEfficiencySummary(
                scope_code=scope_code,
                month_start=month_start,
                planned_output_mt=payload["planned_output_mt"],
                actual_output_mt=payload["actual_output_mt"],
                efficiency_ratio=payload["efficiency_ratio"],
            )
        )
        imported += 1

    return imported


def run_raw_data_import():
    plant_map, machine_map = ensure_master_data()
    clear_raw_tables()

    product_cache = {}
    daily_count = import_daily_product_tonnage(RAW_FILES["daily"], plant_map, product_cache)
    monthly_product_count = import_monthly_product_tonnage(RAW_FILES["monthly"], plant_map, product_cache)
    monthly_machine_count = import_monthly_machine_summary(RAW_FILES["monthly"], machine_map)
    daily_plan_count = import_daily_machine_plan(RAW_FILES["plan"], machine_map)
    efficiency_count = import_monthly_efficiency_summary(RAW_FILES["plan"])

    db.session.commit()

    return {
        "products": len(product_cache),
        "daily_product_tonnage_rows": daily_count,
        "monthly_product_tonnage_rows": monthly_product_count,
        "daily_machine_plan_rows": daily_plan_count,
        "monthly_machine_summary_rows": monthly_machine_count,
        "monthly_efficiency_summary_rows": efficiency_count,
    }


def import_raw_data():
    from backend import create_app

    for name, path in RAW_FILES.items():
        if not path.exists():
            raise FileNotFoundError(f"Missing {name} workbook: {path}")

    app = create_app("development")
    with app.app_context():
        summary = run_raw_data_import()

        print("Raw data import completed.")
        print(f"Products: {summary['products']}")
        print(f"Daily product tonnage rows: {summary['daily_product_tonnage_rows']}")
        print(f"Monthly product tonnage rows: {summary['monthly_product_tonnage_rows']}")
        print(f"Daily machine plan rows: {summary['daily_machine_plan_rows']}")
        print(f"Monthly machine summary rows: {summary['monthly_machine_summary_rows']}")
        print(f"Monthly efficiency summary rows: {summary['monthly_efficiency_summary_rows']}")


if __name__ == "__main__":
    import_raw_data()
